# -*- coding:utf-8 -*-
# @Time  : 2019/4/15 18:17
# @Author: xiaoxiao
# @File  : do_Excel.py

from openpyxl import load_workbook

class Case:
    def __init__(self):
        self.caseid=None
        self.title=None
        self.url=None
        self.data=None
        self.method=None
        self.expected=None
        self.actual=None
        self.result=None

class do_excel():
    def __init__(self,filename,sheetname):
        self.filename=filename
        self.workbook=load_workbook(self.filename)
        self.sheetname = self.workbook[sheetname]
    def read(self):
        row=self.sheetname.max_row
        cases = []
        for row in range(2,row+1):
            case = Case()
            case.caseid=self.sheetname.cell(row,1).value
            case.title=self.sheetname.cell(row,2).value
            case.url = self.sheetname.cell(row, 3).value
            case.data = self.sheetname.cell(row, 4).value
            case.method = self.sheetname.cell(row, 5).value
            case.expected = self.sheetname.cell(row, 6).value
            cases.append(case)
        return cases

    def write(self,str,row,column):
        self.sheetname.cell(row,column).value=str

    def close(self):
        self.workbook.save(self.filename)
        self.workbook.close()



