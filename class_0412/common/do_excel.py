# -*- coding:utf-8 -*-
# @Time  : 2019/4/13 22:37
# @Author: xiaoxiao
# @File  : do_excel.py

from openpyxl import load_workbook


class do_excel:
    def __init__(self,filename,sheetname):
        self.filename=filename
        self.workbook=load_workbook(self.filename)
        self.sheetname = self.workbook[sheetname]
    def read(self,row=None,column=None):
        if not row:
            row=self.sheetname.max_row
        if not column:
            column=self.sheetname.max_column
        cases=[]
        for row in range(1,row+1):
            case=[]
            for column in range(1,column+1):
                value=self.sheetname.cell(row+1,column).value
                if value:
                    case.append(value)
            if case:
                cases.append(case)
        return cases
    def write(self,str,row,column):
        self.sheetname.cell(row,column).value=str

    def close(self):
        self.workbook.save(self.filename)
        self.workbook.close()

if __name__ == '__main__':
    # from class_0412.common.filename import filename
    # s= do_excel(filename,"login").read()
    # print(s)
    import math
    print(2**15)




