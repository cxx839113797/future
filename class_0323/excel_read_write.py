# -*- coding:utf-8 -*-
# @Time  : 2019/3/14 23:50
# @Author: xiaoxiao
# @File  : excel_read_write.py

from openpyxl import workbook
from openpyxl import load_workbook

class excel_read_write:
    def __init__(self,filename,sheet_name):
        self.filename=filename
        self.sheet_name=sheet_name
    def excel_read(self):
        try:
            wb=load_workbook(self.filename)
        except FileNotFoundError as e:
            print("文件不存在！！")
        else:
            if self.sheet_name in wb.sheetnames:
                sheet=wb[self.sheet_name]
                L = []
                for i in range(2,sheet.max_row+1):
                    value=[]
                    for j in range(2,sheet.max_column+1):
                        value.append(sheet.cell(i, j).value)
                    L.append(value)
                wb.close()
                return L
    def excel_write(self,str,row,column):
        try:
            wb=load_workbook(self.filename)
            if self.sheet_name not in wb.sheetnames:
                wb.create_sheet(self.sheet_name)
        except FileNotFoundError as e:
            wb=workbook.Workbook()
            wb.create_sheet(self.sheet_name)
        finally:
            sheet=wb[self.sheet_name]
            sheet.cell(row,column).value=str
            wb.save(self.filename)
            wb.close()


if __name__ == '__main__':
    print(excel_read_write("TestData.xlsx", "test_data").excel_read())





